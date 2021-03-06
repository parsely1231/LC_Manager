from src.models.hplc_data import ExperimentalData, DataTable, ImpurityData

import openpyxl as excel
from openpyxl.styles import PatternFill


class ExcelModel:
    TABLE_SIZE = 7
    ELEMENT_POS = {
        'name': 0,
        'RT': 1,
        'RRT': 2,
        'Area': 3,
        'Area%': 4,
        '補正Area%': 5,
    }
    INIT_ROW = 2
    INIT_COL = 2

    def __init__(self):
        self.wb = excel.Workbook()
        self.ws = self.wb.active

    def to_xlsx(self, exp: ExperimentalData, file_path):
        x = self.INIT_COL
        y = self.INIT_ROW
        for _, table in exp.tables.items():
            self.write_table(table, y, x, exp.imp_excluded)
            x += self.TABLE_SIZE

        self.write_imp_trend(exp)
        self.wb.save(file_path)
        self.wb.close()

    def write_table(self, table: DataTable, y, x, excluded):
        """worksheetにtable一つ分のデータを書き込む"""

        def write_header():
            self.ws.cell(y, x, table.name)  # table.nameを初期位置に書く
            for element, dx in self.ELEMENT_POS.items():  # テーブルのカラム名を初期位置の1段下に書く
                nx = x + dx
                self.ws.cell(y+1, nx, element)

        def write_footer():
            final_row = y + 2 + len(table.imp_list)
            self.ws.cell(final_row, x + self.ELEMENT_POS['Area'] - 1, 'Total Area')
            self.ws.cell(final_row + 1, x + self.ELEMENT_POS['Area'] - 1, 'Excluded')
            self.ws.cell(final_row, x + self.ELEMENT_POS['Area'], table.total_area)
            self.ws.cell(final_row + 1, x + self.ELEMENT_POS['Area'], table.edited_total_area)

        def paint_cells(py, px, color):
            for col in range(px, px + self.TABLE_SIZE - 1):
                self.ws.cell(py, col).fill = color

        write_header()
        gray = 'd3d3d3'
        cell_color = PatternFill(fill_type='solid', fgColor=gray)

        for dy, imp_data in enumerate(table.imp_list, 2):  # 各不純物のデータを書く
            self.write_imp_data(imp_data, y+dy, x)
            if imp_data.name in excluded:
                paint_cells(y+dy, x, cell_color)

        write_footer()

    def write_imp_data(self, data: ImpurityData, y, x):
        self.ws.cell(y, x+self.ELEMENT_POS['name'], data.name)
        self.ws.cell(y, x+self.ELEMENT_POS['RT'], data.rt)
        self.ws.cell(y, x+self.ELEMENT_POS['RRT'], data.rrt)
        self.ws.cell(y, x+self.ELEMENT_POS['Area'], data.area)
        self.ws.cell(y, x+self.ELEMENT_POS['Area%'], data.area_ratio)
        self.ws.cell(y, x+self.ELEMENT_POS['補正Area%'], data.edited_area_ratio)

    def write_imp_trend(self, exp: ExperimentalData):
        self.ws = self.wb.create_sheet('Impurity Trend')
        init_x = 2
        init_y = 2
        rrt_to_row = {}
        rrt_list = sorted(list(exp.rrt_set))
        for row, rrt in enumerate(rrt_list, init_y+1):
            rrt_to_row[rrt] = row
            self.ws.cell(row, init_x, rrt)

        for col, (name, table) in enumerate(exp.tables.items(), init_x+1):
            self.ws.cell(init_y, col, name)
            for imp in table.imp_list:
                row = rrt_to_row[imp.rrt]
                self.ws.cell(row, col, imp.edited_area_ratio)
